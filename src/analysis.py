import os
import glob
import openpyxl
import sympy as sp
import matplotlib.pyplot as plt

# --- KONFIGURACJA ŚCIEŻEK ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(BASE_DIR)

DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
IMG_DIR = os.path.join(PROJECT_ROOT, 'images')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(IMG_DIR, exist_ok=True)

PLIK_BAZA = os.path.join(DATA_DIR, 'materialy_budowlane.xlsx')

# Warunki brzegowe
T_WEW = 20.0   
T_ZEW = -20.0  
R_SI = 0.13    
R_SE = 0.04    

# --- CZĘŚĆ 1: Obsługa Excela ---
def przygotuj_baze_materialow(sciezka_do_bazy):
    if not os.path.exists(sciezka_do_bazy):
        print(f"[INFO] Tworzę nową bazę materiałów: {sciezka_do_bazy}...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materialy"
        ws.append(["Nazwa", "Lambda [W/mK]", "Opis"])
        dane = [
            ("Beton zwykły", 1.70, "Konstrukcyjny"),
            ("Pustak ceramiczny", 0.30, "Porotherm itp."),
            ("Cegła pełna", 0.77, "Klasyczna"),
            ("Styropian EPS", 0.040, "Izolacja standard"),
            ("Styropian Grafitowy", 0.031, "Izolacja premium"),
            ("Wełna mineralna", 0.035, "Izolacja akustyczna/ognio"),
            ("Tynk cementowo-wapienny", 0.82, "Wykończenie wew."),
            ("Tynk cienkowarstwowy", 1.00, "Elewacja")
        ]
        for wiersz in dane:
            ws.append(wiersz)
        wb.save(sciezka_do_bazy)
        print("[INFO] Baza gotowa.")
    else:
        print(f"[INFO] Znaleziono bazę: {os.path.basename(sciezka_do_bazy)}")

def pobierz_lambde(nazwa_materialu, sciezka_do_bazy):
    wb = openpyxl.load_workbook(sciezka_do_bazy, data_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat_nazwa = row[0].strip()
        mat_lambda = row[1]
        if mat_nazwa == nazwa_materialu:
            return float(mat_lambda)
            
    raise ValueError(f"Nie znaleziono materiału '{nazwa_materialu}' w bazie!")

# --- CZĘŚĆ 2: Obliczenia Symboliczne SymPy ---
def oblicz_przegrode_symbolicznie(warstwy_dane):
    # (Bez zmian w logice obliczeń)
    R_tot_sym = sp.symbols('R_tot')
    R_si = sp.symbols('R_si')
    R_se = sp.symbols('R_se')
    
    n = len(warstwy_dane)
    d_syms = sp.symbols(f'd_1:{n+1}')       
    lam_syms = sp.symbols(f'lambda_1:{n+1}') 
    
    suma_oporow = sum(d/l for d, l in zip(d_syms, lam_syms))
    rownanie_R = sp.Eq(R_tot_sym, R_si + suma_oporow + R_se)
    
    wartosci = {R_si: R_SI, R_se: R_SE}
    for i, (nazwa, d_val, lam_val) in enumerate(warstwy_dane):
        wartosci[d_syms[i]] = d_val
        wartosci[lam_syms[i]] = lam_val
        
    R_wynik = rownanie_R.rhs.subs(wartosci)
    U_wynik = 1 / R_wynik
    
    q = U_wynik * (T_WEW - T_ZEW)
    temperatury = [T_WEW]
    aktualna_T = T_WEW
    
    t_surf_in = aktualna_T - q * R_SI
    temperatury.append(t_surf_in)
    aktualna_T = t_surf_in
    
    x_coords = [0]
    current_x = 0
    
    for _, d, lam in warstwy_dane:
        delta_T_warstwy = q * (d / lam)
        aktualna_T -= delta_T_warstwy
        temperatury.append(aktualna_T)
        current_x += d
        x_coords.append(current_x)
        
    temperatury.append(T_ZEW)
    
    return float(U_wynik), float(R_wynik), x_coords, temperatury

# --- CZĘŚĆ 3: Wizualizacja ---
def rysuj_wykres(x_coords, temperatury, warstwy_dane, u_val, nazwa_pliku_wykresu):    
    x_plot = [-0.05] + x_coords + [x_coords[-1] + 0.05] 
    y_plot = temperatury
    
    plt.figure(figsize=(10, 6))
    plt.plot(x_plot, y_plot, 'r-o', linewidth=3, label='Temp [°C]')
    
    colors = ['#f0f0f0', '#ffcc99', '#e6f2ff', '#d9d9d9', '#ffe6cc']
    prev_x = 0
    
    for i, (nazwa, d, lam) in enumerate(warstwy_dane):
        color = colors[i % len(colors)]
        plt.axvspan(prev_x, prev_x + d, color=color, alpha=0.5, ymin=0, ymax=1)
        plt.text(prev_x + d/2, min(temperatury)+5, f"{nazwa}\n{d}m", 
                 rotation=90, ha='center', va='bottom', fontsize=8)
        prev_x += d

    plt.axhline(0, color='black', linewidth=0.8, linestyle='--')
    plt.title(f'Wynik: {os.path.basename(nazwa_pliku_wykresu)} (U = {u_val:.3f})', fontsize=12)
    plt.xlabel('Grubość [m]')
    plt.grid(True, linestyle='--')
    plt.legend()
    
    sciezka_zapisu = os.path.join(IMG_DIR, nazwa_pliku_wykresu)
    plt.savefig(sciezka_zapisu, dpi=150)
    plt.close()
    print(f"[SUKCES] Wygenerowano wykres: {sciezka_zapisu}")
    
def wczytaj_projekt_z_excela(sciezka_pliku):
    wb = openpyxl.load_workbook(sciezka_pliku, data_only=True)
    ws = wb.active
    sciana = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            sciana.append((row[0], float(row[1])))
    return sciana
    
# --- MAIN ---
if __name__ == "__main__":
    print("=== U-VALUE SOLVER: TRYB WSADOWY ===")
    print(f"[INFO] Katalog roboczy (src): {BASE_DIR}")
    print(f"[INFO] Dane wejściowe (data): {DATA_DIR}")
    print(f"[INFO] Wyniki (images): {IMG_DIR}")
    
    przygotuj_baze_materialow(PLIK_BAZA)
    
    szablon_wyszukiwania = os.path.join(DATA_DIR, "*.xlsx")
    pliki_wszystkie = glob.glob(szablon_wyszukiwania)
    
    pliki_projektowe = [f for f in pliki_wszystkie if os.path.basename(f) != 'materialy_budowlane.xlsx']
    
    if not pliki_projektowe:
        print("[INFO] Brak plików projektowych w folderze 'data'. Uruchom najpierw generator!")
    else:
        print(f"Znaleziono {len(pliki_projektowe)} projektów do policzenia.")
    
    for pelna_sciezka in pliki_projektowe:
        nazwa_pliku = os.path.basename(pelna_sciezka)
        print(f"\n---> Przetwarzam: {nazwa_pliku}")
        try:
            warstwy_z_projektu = wczytaj_projekt_z_excela(pelna_sciezka)
            
            sciana_do_obliczen = []
            for nazwa, grubosc in warstwy_z_projektu:
                lam = pobierz_lambde(nazwa, PLIK_BAZA)
                sciana_do_obliczen.append((nazwa, grubosc, lam))
            
            u_val, r_val, x_c, temps = oblicz_przegrode_symbolicznie(sciana_do_obliczen)
            
            nazwa_wykresu = nazwa_pliku.replace(".xlsx", " ")
            rysuj_wykres(x_c, temps, sciana_do_obliczen, u_val, nazwa_wykresu)
            
            if u_val < 0.20:
                print(f"     STATUS: OK (U={u_val:.3f})")
            else:
                print(f"     STATUS: ZA SŁABA IZOLACJA! (U={u_val:.3f})")
                
        except Exception as e:
            print(f"     [BŁĄD] Plik {nazwa_pliku}: {e}")
            
    print("\n=== KONIEC ===")